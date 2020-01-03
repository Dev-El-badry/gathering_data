from bs4 import BeautifulSoup
import requests
from oop.Single_Item import Single_Item
from oop.OthrItem import OthrItem
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, NamedStyle, PatternFill
import time, os, re, math


class Project(Single_Item, OthrItem):

    #empty data array
    data_result = []

    # Declartion Varables show in console
    const_pagination_num = 0
    pagination_num_now = 0
    const_items_num = 0
    items_num_now_download = 0
    items_num_now_not_download = 0
    
    #empty  list
    __items_list = []
    empty_list = []
    #Get Unixsttamp
    unix_dt = int(time.time())

    wb = Workbook()
    sh = wb.active

    def __init__(self, seller_name):
        self.seller_name = seller_name

        #start work get html
        self.start_main_function()
    
    #GET data from pagination and lazy loadin 
    #pgaintor range (1,52)
    #max number: 3120



    def __get_target_url(self, num=1):
        target_url = "https://egypt.souq.com/eg-ar/" + \
            self.seller_name+"/s?section=2&page=" + str(num)
        return target_url

    def get_data_from_single_item(self, url) :
        
        single_item = Single_Item(url)
        data = single_item.getDataFromURL()
        return data

    def get_othr_name_of_seller(self, url) :
        othr_seller = OthrItem(url, self.seller_name)
        data = othr_seller.get_data_belong_to_seller_name()
        return data

    def setup_excel(self) :
        #setup excel
        
        #self.sh.sheet_view.rightToLeft=True
        self.sh.title = "Report"
        self.wb.create_sheet(title="items")
        self.sh.sheet_properties.tabColor = "1072BA"

        # ##Let's create a style template for the header row
        # header_line1 = NamedStyle(name="header")
        # header_line1.font = Font(bold=True)
        # header_line1.border = Border(bottom=Side(border_style="thin"))
        # header_line1.alignment = Alignment(horizontal="center", vertical="center")

        # pattern fill for cell
        # redFill = PatternFill(start_color='FFFF0000',
        #                 end_color='FFFF0000',
        #                 fill_type='solid')

        #  list_data = {
                #     'item_title': item_title.get_text(),
                #     'url_item': url_item,
                #     'item_price': item_price.get_text(),
                #     'item_seller_name': seller_of_item_name,
                #     'ur_name': ur_name,
                #     'ur_offer': ur_offer
                # }

        headers = ['Item ID', 'Item Title', 'href-url', 'Item Price','Competitor Seller', 'Your Name', 'Your Price OFfer']

       

        c = 1
        for i in headers:
            self.sh.cell(row=1, column=c).value = i
            c += 1

        ##Now let's apply this to all first row (header) cells
        # header_row = self.sh[1]
        # for cell in header_row:
        #     cell.style = header_line1

        return

    def get_count_pagination(self, num) :
        res = int(num) / 60
        return math.ceil(res)

    def get_numbers_from_text(self, string):
        str = string
        x = re.findall(r'\d+', string)
        nums = ''.join(x)
        return nums
    
    def start_main_function(self) :
        const_number = 3060
        lst_data = []
        target_url = self.__get_target_url()
        response = requests.get(target_url)
        soup = BeautifulSoup(response.text, 'lxml')

        #Some Usfull Data
        num_items = soup.find(class_="top").find(class_="total").get_text()

        number = self.get_numbers_from_text(num_items)
        if int(number) < const_number:
            num_pagination = self.get_count_pagination(number)
            self.const_items_num = number
        else:
            num_pagination = 51
            self.const_items_num = const_number

        self.const_pagination_num = num_pagination
        

        
        #call setup excel
        self.setup_excel()
        for x in range(1,num_pagination+1) :
            self.pagination_num_now = x
            new_target_url = self.__get_target_url(x)
            result = self.main_function(new_target_url)

        i=1
        r=2
        self.start_write_in_excel2(i,r)

        self.save_file_excel()

        

        return self.empty_list
       


    def main_function(self, target_url) :
        #declration local varables 
        ur_name = ''
        ur_offer = ''
        
        
        try:
            response = requests.get(target_url)
        except requests.exceptions.ConnectionError:
            r.status_code = "Connection refused"
        soup = BeautifulSoup(response.text, 'lxml')
       
        items = soup.find_all(class_="single-item")

      
       
        count = 0
        count_x= 1
        count_y= 1
        list_data = {}
        for item in range(len(items)):
            
           
            itemID = items[item].find().parent.attrs['data-ean']
            
            item_title = items[item].find(class_='itemTitle')
            item_price = items[item].find(class_='itemPrice')
            url_item = items[item].find(class_="quickViewAction")['href']
            new_url = url_item.replace(url_item[-2], 'io')
            # priceItem = self.get_othr_price_of_item(new_url)
            
            #THE SELLER SHOW IN SINGLE ITEM (SHOW IN SEARCH)
            single_item_data = self.get_data_from_single_item(url_item)
            
            if(single_item_data['seller_name'] != self.seller_name):
                
                self.items_num_now_download = self.items_num_now_download + 1

                res_data = self.get_othr_name_of_seller(new_url)
                ur_name = res_data['seller_name']
                ur_offer = res_data['offer_price']

                # list = [list_data['nw_id_item'], list_data['item_title'], list_data['url_item'],
                # list_data['item_price'], list_data['item_seller_name'], list_data['ur_name'], list_data['ur_offer']]

                list_data = {
                    'nw_id_item': itemID,

                    'item_title': item_title.get_text(),
                    'url_item': url_item,
                    'item_price': item_price.get_text(),
                    'item_seller_name': single_item_data['seller_name'],
                    'ur_name': ur_name,
                    'ur_offer': ur_offer
                    
                }

                self.data_result.append(list_data)
                

               
            else :
                self.items_num_now_not_download = self.items_num_now_not_download + 1

            print('===========')
            print("pagination = {} / {} ".format(self.const_pagination_num, self.pagination_num_now))
            print('=============')


            print('===========')
            print("items = not show => {} / show => {} / from => {}  ".format(self.items_num_now_download,
                                            self.items_num_now_not_download, self.const_items_num))
            print('=============')


      
        
    def start_write_in_excel2(self, i,r) :
        print('start write in excel')

        for list in self.data_result :
            for lst in list :
                if i > len(list):
                    i = 1

                self.sh.cell(row=r, column=i).value = list[lst]
                i +=1 
            r+=1
        
        print('end write in excel')

    def start_write_in_excel(self, i, r, list_data) :
        list = [list_data['nw_id_item'], list_data['item_title'], list_data['url_item'],
                list_data['item_price'], list_data['item_seller_name'], list_data['ur_name'], list_data['ur_offer']]
        print('start')
        for lst1 in list:
            if i > len(list):
                i = 1

            self.sh.cell(row=r, column=i).value = lst1
            i += 1
        print('end')

    def save_file_excel(self) :
       
        filename = 'Report-'+str(self.unix_dt)+'.xlsx'
        home_path = os.path.join(os.environ['HOMEPATH'], 'Desktop')
        print('Finish')
        #Saving File
        self.wb.save(os.path.join(home_path, filename))









































